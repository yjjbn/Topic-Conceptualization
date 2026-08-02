"""Combine model initial-classification JSON files into one Excel workbook.

By default, this script reads the Gemma, Llama, and GPT JSON files from the
``out`` directory beside this script and writes
``out/combined_initial_classifications.xlsx``.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


MODEL_PATTERNS = {
    "gemma": "*gemma*initial_classifications.json",
    "llama": "*llama*initial_classifications.json",
    "gpt": "*gpt*initial_classifications.json",
}

HEADERS = [
    "document_id",
    "text",
    "label_gemma",
    "decision_basis_gemma",
    "label_llama",
    "decision_basis_llama",
    "label_gpt",
    "decision_basis_gpt",
    "error_llama",
    "error_gpt",
]


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Combine Gemma, Llama, and GPT classifications into Excel."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=script_dir / "out",
        help="Directory containing the JSON inputs (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "Output workbook path (default: "
            "<input-dir>/combined_initial_classifications.xlsx)"
        ),
    )
    return parser.parse_args()


def find_source(input_dir: Path, model: str) -> Path:
    matches = sorted(input_dir.glob(MODEL_PATTERNS[model]))
    if len(matches) != 1:
        names = ", ".join(path.name for path in matches) or "none"
        raise RuntimeError(
            f"Expected exactly one {model} input matching "
            f"{MODEL_PATTERNS[model]!r}; found {len(matches)}: {names}"
        )
    return matches[0]


def load_records(path: Path) -> tuple[list[str], dict[str, dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig") as file:
        raw = json.load(file)

    if not isinstance(raw, list):
        raise ValueError(f"{path.name}: JSON root must be an array")

    order: list[str] = []
    records: dict[str, dict[str, Any]] = {}
    required = {"document_id", "text", "label", "decision_basis"}
    for index, record in enumerate(raw, start=1):
        if not isinstance(record, dict):
            raise ValueError(f"{path.name}: item {index} is not an object")
        missing = required.difference(record)
        if missing:
            raise ValueError(
                f"{path.name}: item {index} is missing {sorted(missing)}"
            )
        document_id = str(record["document_id"])
        if document_id in records:
            raise ValueError(f"{path.name}: duplicate document_id {document_id!r}")
        order.append(document_id)
        records[document_id] = record
    return order, records


def same_label(baseline: dict[str, Any] | None, other: dict[str, Any] | None) -> int:
    """Return 1 for an exact label match with Gemma, otherwise return 0."""
    if baseline is None or other is None:
        return 0
    return int(other["label"] == baseline["label"])


def build_workbook(input_dir: Path, output_path: Path) -> tuple[int, dict[str, Path]]:
    sources = {model: find_source(input_dir, model) for model in MODEL_PATTERNS}
    orders: dict[str, list[str]] = {}
    data: dict[str, dict[str, dict[str, Any]]] = {}
    for model, path in sources.items():
        orders[model], data[model] = load_records(path)

    document_ids: list[str] = []
    seen: set[str] = set()
    for model in ("gemma", "llama", "gpt"):
        for document_id in orders[model]:
            if document_id not in seen:
                seen.add(document_id)
                document_ids.append(document_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "initial_classifications"
    sheet.append(HEADERS)

    for document_id in document_ids:
        records = {model: data[model].get(document_id) for model in MODEL_PATTERNS}
        texts = {
            record["text"]
            for record in records.values()
            if record is not None
        }
        if len(texts) > 1:
            raise ValueError(
                f"Conflicting text values for document_id {document_id!r}"
            )
        text = next(iter(texts), "")
        row: list[Any] = [document_id, text]
        for model in ("gemma", "llama", "gpt"):
            record = records[model]
            row.extend(
                [record["label"], record["decision_basis"]]
                if record is not None
                else ["", ""]
            )
        row.extend(
            [
                same_label(records["gemma"], records["llama"]),
                same_label(records["gemma"], records["gpt"]),
            ]
        )
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Prevent strings beginning with '=' from becoming Excel formulas.
            if isinstance(cell.value, str):
                cell.data_type = "s"

    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.zoomScale = 85
    widths = [38, 85, 24, 60, 24, 60, 24, 60, 14, 14]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    sheet.row_dimensions[1].height = 24

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(document_ids), sources


def main() -> None:
    args = parse_args()
    input_dir = args.input_dir.resolve()
    output_path = (
        args.output.resolve()
        if args.output is not None
        else input_dir / "combined_initial_classifications.xlsx"
    )
    row_count, sources = build_workbook(input_dir, output_path)
    for model, path in sources.items():
        print(f"{model}: {path.name}")
    print(f"Wrote {row_count} rows to {output_path}")


if __name__ == "__main__":
    main()
